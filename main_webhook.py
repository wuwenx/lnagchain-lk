"""
飞书 Webhook（请求地址）模式：HTTP 服务接收事件 → LangChain 回复 → 飞书 API 发回
在 Lark 后台「事件配置」里将「请求地址」设为：https://你的公网域名/ 或 /webhook 等（与 WEBHOOK_PATH 一致）
"""
import os
import certifi
os.environ.setdefault("SSL_CERT_FILE", certifi.where())
os.environ.setdefault("REQUESTS_CA_BUNDLE", certifi.where())

import html
import logging
from contextlib import asynccontextmanager

from fastapi import FastAPI, Form, HTTPException, Request, Response
from lark_oapi.core.exception import EventException
from lark_oapi.core.model import RawRequest, RawResponse
from apscheduler.schedulers.background import BackgroundScheduler

from config import (
    FEISHU_BINANCE_ANNOUNCEMENTS_CHAT_ID,
    FEISHU_BYBIT_ANNOUNCEMENTS_CHAT_ID,
    FEISHU_ENCRYPT_KEY,
    FEISHU_MEXC_DELISTINGS_CHAT_ID,
    FEISHU_NEEDLE_ALERT_CHAT_ID,
    FEISHU_OKX_ANNOUNCEMENTS_CHAT_ID,
    FEISHU_POPFUN_LOG_CHAT_ID,
    FEISHU_TOOBIT_24H_CHAT_ID,
    FEISHU_VERIFICATION_TOKEN,
    WEBHOOK_HOST,
    WEBHOOK_PATH,
    WEBHOOK_PORT,
    validate_webhook_config,
    update_runtime_config_file,
)
from gitlab_webhook import handle_gitlab_webhook
from handlers import build_event_handler
from tasks.binance_announcements import run_binance_announcements_push
from tasks.bybit_announcements import run_bybit_announcements_push
from tasks.mexc_delistings import run_mexc_delistings_push
from tasks.needle_scan import run_needle_scan_push
from tasks.okx_announcements import run_okx_announcements_push
from tasks.popfun_log import run_popfun_log_push
from tasks.toobit_24h import run_toobit_24h_push

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

_scheduler: BackgroundScheduler | None = None
_ADMIN_PAGE_TOKEN = (os.environ.get("ADMIN_PAGE_TOKEN", "") or "").strip()


def _check_admin_token(request: Request, token: str | None = None) -> None:
    """可选鉴权：若设置 ADMIN_PAGE_TOKEN，则需在 query/form/header 传入相同 token。"""
    if not _ADMIN_PAGE_TOKEN:
        return
    q = (request.query_params.get("token", "") or "").strip()
    f = (token or "").strip()
    h = (request.headers.get("X-Admin-Token", "") or "").strip()
    if not (q == _ADMIN_PAGE_TOKEN or f == _ADMIN_PAGE_TOKEN or h == _ADMIN_PAGE_TOKEN):
        raise HTTPException(status_code=403, detail="forbidden")


def _render_model_admin_page(message: str = "") -> str:
    from config import (
        OPENAI_API_BASE,
        OPENAI_API_KEY,
        OPENAI_MODEL,
        VISION_MULTIMODAL,
        VISION_OPENAI_API_BASE,
        VISION_OPENAI_API_KEY,
        VISION_OPENAI_MODEL,
    )

    api_base = OPENAI_API_BASE or ""
    api_key = OPENAI_API_KEY or ""
    model = OPENAI_MODEL or ""
    v_base = VISION_OPENAI_API_BASE or ""
    v_key = VISION_OPENAI_API_KEY or ""
    v_model = VISION_OPENAI_MODEL or ""
    key_masked = (
        (api_key[:8] + "..." + api_key[-4:]) if len(api_key) > 14 else ("*" * len(api_key) if api_key else "(未设置)")
    )
    v_key_masked = (
        (v_key[:8] + "..." + v_key[-4:]) if len(v_key) > 14 else ("*" * len(v_key) if v_key else "(未设置)")
    )
    sel_off = "" if VISION_MULTIMODAL else " selected"
    sel_on = " selected" if VISION_MULTIMODAL else ""
    msg_html = (
        f'<p style="padding:10px;border-radius:6px;background:#f0fdf4;color:#166534;">{html.escape(message)}</p>'
        if message
        else ""
    )
    token_tip = (
        "<p><strong>鉴权：</strong>已启用 ADMIN_PAGE_TOKEN，请在 URL 加 <code>?token=xxx</code>，"
        "或 POST 表单中带 <code>token</code>。</p>"
        if _ADMIN_PAGE_TOKEN
        else "<p><strong>鉴权：</strong>未启用（建议配置环境变量 <code>ADMIN_PAGE_TOKEN</code>）。</p>"
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>模型配置管理</title>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif; margin: 24px; color: #111827; }}
    .wrap {{ max-width: 760px; margin: 0 auto; }}
    label {{ display: block; margin: 14px 0 6px; font-weight: 600; }}
    input, select {{ width: 100%; padding: 10px; border: 1px solid #d1d5db; border-radius: 8px; }}
    h3 {{ margin-top: 22px; font-size: 15px; color: #374151; }}
    button {{ margin-top: 16px; padding: 10px 14px; border: 0; border-radius: 8px; background: #2563eb; color: white; cursor: pointer; }}
    code {{ background: #f3f4f6; padding: 2px 6px; border-radius: 4px; }}
    .hint {{ color: #4b5563; font-size: 14px; }}
  </style>
</head>
<body>
  <div class="wrap">
    <h2>模型配置管理</h2>
    <p class="hint"><strong>LLM 模型相关仅从项目根 <code>config.json</code> 读取，不使用 <code>.env</code>。</strong>保存后写入该文件（可复制 <code>config.example.json</code>）。</p>
    <p class="hint">需识别飞书/Lark <strong>聊天截图或文档内嵌图</strong>时：将 <code>VISION_MULTIMODAL</code> 设为开启，并填写下方多模态三套；含图请求将<strong>只走多模态接口</strong>，不再使用文本模型的 <code>OPENAI_MODEL</code>。</p>
    <p class="hint">文本 Key：<code>{html.escape(key_masked)}</code>；多模态 Key：<code>{html.escape(v_key_masked)}</code>（留空则不修改对应 Key）。</p>
    {token_tip}
    {msg_html}
    <form method="post" action="/admin/model">
      <h3>文本对话（OpenAI 兼容）</h3>
      <label>OPENAI_API_BASE（接口地址）</label>
      <input name="openai_api_base" value="{html.escape(api_base)}" placeholder="例如 https://api.openai.com/v1 或 https://api.deepseek.com/v1" required />

      <label>OPENAI_API_KEY（留空则保持不变）</label>
      <input type="password" name="openai_api_key" value="" placeholder="输入新 key；留空表示不修改" />

      <label>OPENAI_MODEL（文本模型）</label>
      <input name="openai_model" value="{html.escape(model)}" placeholder="例如 gpt-4o-mini / deepseek-chat" required />

      <h3>多模态（Lark 截图 / 图片）</h3>
      <label>VISION_MULTIMODAL</label>
      <select name="vision_multimodal">
        <option value="false"{sel_off}>关闭（含图消息将丢弃图片，仅文本模型）</option>
        <option value="true"{sel_on}>开启（含图时仅使用下方多模态配置，忽略文本模型名）</option>
      </select>

      <label>VISION_OPENAI_API_BASE</label>
      <input name="vision_openai_api_base" value="{html.escape(v_base)}" placeholder="可与文本相同或单独的多模态端点" />

      <label>VISION_OPENAI_API_KEY（留空则保持不变）</label>
      <input type="password" name="vision_openai_api_key" value="" placeholder="输入新 key；留空表示不修改" />

      <label>VISION_OPENAI_MODEL</label>
      <input name="vision_openai_model" value="{html.escape(v_model)}" placeholder="例如 gpt-4o" />

      <label>token（可选，仅当启用 ADMIN_PAGE_TOKEN）</label>
      <input name="token" value="" placeholder="与 ADMIN_PAGE_TOKEN 一致" />

      <button type="submit">保存到 config.json</button>
    </form>
    <p class="hint" style="margin-top: 16px;">保存后会重新加载配置并清空简单对话链缓存；若仍有异常可重启 <code>main_webhook.py</code>。</p>
  </div>
</body>
</html>"""


@asynccontextmanager
async def _lifespan(app: FastAPI):
    """应用生命周期：启动时启动定时任务，关闭时停止。"""
    global _scheduler
    has_toobit = bool((FEISHU_TOOBIT_24H_CHAT_ID or "").strip())
    has_needle = bool((FEISHU_NEEDLE_ALERT_CHAT_ID or "").strip())
    has_mexc = bool((FEISHU_MEXC_DELISTINGS_CHAT_ID or "").strip())
    has_binance = bool((FEISHU_BINANCE_ANNOUNCEMENTS_CHAT_ID or "").strip())
    has_okx = bool((FEISHU_OKX_ANNOUNCEMENTS_CHAT_ID or "").strip())
    has_bybit = bool((FEISHU_BYBIT_ANNOUNCEMENTS_CHAT_ID or "").strip())
    has_popfun = bool((FEISHU_POPFUN_LOG_CHAT_ID or "").strip())
    if has_toobit or has_needle or has_mexc or has_binance or has_okx or has_bybit or has_popfun:
        _scheduler = BackgroundScheduler()
        # if has_toobit:
        #     _scheduler.add_job(run_toobit_24h_push, "interval", minutes=5, id="toobit_24h")
        #     logger.info("Toobit 24h scheduler (every 5 min -> %s)", (FEISHU_TOOBIT_24H_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_toobit_24h_push()
        #     except Exception as e:
        #         logger.exception("Toobit 24h first run error: %s", e)
        # if has_needle:
        #     _scheduler.add_job(run_needle_scan_push, "interval", minutes=5, id="needle_scan")
        #     logger.info("Needle scan scheduler (every 5 min -> %s)", (FEISHU_NEEDLE_ALERT_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_needle_scan_push()
        #     except Exception as e:
        #         logger.exception("Needle scan first run error: %s", e)
        # if has_mexc:
        #     _scheduler.add_job(run_mexc_delistings_push, "interval", minutes=5, id="mexc_delistings")
        #     logger.info("MEXC delistings scheduler (every 5 min -> %s)", (FEISHU_MEXC_DELISTINGS_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_mexc_delistings_push()
        #     except Exception as e:
        #         logger.exception("MEXC delistings first run error: %s", e)
        # if has_binance:
        #     _scheduler.add_job(run_binance_announcements_push, "interval", minutes=5, id="binance_announcements")
        #     logger.info("Binance announcements scheduler (every 5 min -> %s)", (FEISHU_BINANCE_ANNOUNCEMENTS_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_binance_announcements_push()
        #     except Exception as e:
        #         logger.exception("Binance announcements first run error: %s", e)
        # if has_okx:
        #     _scheduler.add_job(run_okx_announcements_push, "interval", minutes=5, id="okx_announcements")
        #     logger.info("OKX announcements scheduler (every 5 min -> %s)", (FEISHU_OKX_ANNOUNCEMENTS_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_okx_announcements_push()
        #     except Exception as e:
        #         logger.exception("OKX announcements first run error: %s", e)
        # if has_bybit:
        #     _scheduler.add_job(run_bybit_announcements_push, "interval", minutes=5, id="bybit_announcements")
        #     logger.info("Bybit announcements scheduler (every 5 min -> %s)", (FEISHU_BYBIT_ANNOUNCEMENTS_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_bybit_announcements_push()
        #     except Exception as e:
        #         logger.exception("Bybit announcements first run error: %s", e)
        # if has_popfun:
        #     _scheduler.add_job(run_popfun_log_push, "interval", minutes=15, id="popfun_log")
        #     logger.info("Popfun log scheduler (every 15 min -> %s)", (FEISHU_POPFUN_LOG_CHAT_ID or "")[:20] + "...")
        #     try:
        #         run_popfun_log_push()
        #     except Exception as e:
        #         logger.exception("Popfun log first run error: %s", e)
        _scheduler.start()
    else:
        logger.debug("No FEISHU_*_CHAT_ID set, schedulers disabled")
    yield
    if _scheduler:
        _scheduler.shutdown(wait=False)
        _scheduler = None
        logger.info("Scheduler stopped")


app = FastAPI(title="LangChain + Lark Webhook", lifespan=_lifespan)
event_handler = None


def _get_handler():
    global event_handler
    if event_handler is None:
        event_handler = build_event_handler(
            FEISHU_ENCRYPT_KEY or "",
            FEISHU_VERIFICATION_TOKEN or "",
        )
    return event_handler


def _raw_request_from_http(request: Request, body: bytes) -> RawRequest:
    req = RawRequest()
    req.uri = str(request.url.path or "/")
    req.body = body
    req.headers = {}
    for k, v in request.headers.items():
        if k.lower().startswith("x-lark") or k.lower() in ("content-type",):
            req.headers[k] = v
    return req


def _response_to_http(resp: RawResponse) -> Response:
    ct = (resp.headers or {}).get("Content-Type", "application/json")
    return Response(
        content=resp.content or b"",
        status_code=resp.status_code or 200,
        headers=dict(resp.headers or {}),
        media_type=ct,
    )


@app.post(WEBHOOK_PATH)
@app.get(WEBHOOK_PATH)
async def lark_webhook(request: Request):
    """接收飞书事件：url_verification 或 im.message.receive_v1 等。"""
    body = await request.body()
    raw_req = _raw_request_from_http(request, body)
    handler = _get_handler()
    try:
        raw_resp = handler.do(raw_req)
    except EventException as e:
        if "processor not found" in str(e):
            logger.debug("unhandled event type, return 200: %s", e)
            return Response(content=b'{"msg":"success"}', status_code=200, media_type="application/json")
        raise
    return _response_to_http(raw_resp)


GITLAB_WEBHOOK_PATH = "/webhook/gitlab"


@app.post(GITLAB_WEBHOOK_PATH)
async def gitlab_webhook(request: Request):
    """
    接收 GitLab Webhook（如 Merge request events）。
    GitLab 配置：Settings → Webhooks → URL = https://你的域名/webhook/gitlab，勾选 Merge request events。
    若配置了 GITLAB_WEBHOOK_SECRET，请在同一页填写相同的 Secret token。
    """
    body = await request.body()
    event = request.headers.get("X-Gitlab-Event", "").strip()
    token = request.headers.get("X-Gitlab-Token", "").strip()
    logger.info("GitLab webhook received: X-Gitlab-Event=%s, body_len=%d", event, len(body))
    handled, err = handle_gitlab_webhook(body, event, token or None)
    if err:
        logger.warning("GitLab webhook rejected: %s", err)
        return Response(content=err, status_code=403)
    return Response(content=b'{"ok":true}', status_code=200, media_type="application/json")


@app.get("/admin/model")
async def admin_model_get(request: Request):
    _check_admin_token(request)
    return Response(content=_render_model_admin_page(), media_type="text/html; charset=utf-8")


@app.post("/admin/model")
async def admin_model_post(
    request: Request,
    openai_api_base: str = Form(...),
    openai_api_key: str = Form(""),
    openai_model: str = Form(...),
    vision_multimodal: str = Form("false"),
    vision_openai_api_base: str = Form(""),
    vision_openai_api_key: str = Form(""),
    vision_openai_model: str = Form(""),
    token: str = Form(""),
):
    _check_admin_token(request, token=token)
    api_base = (openai_api_base or "").strip()
    api_key = (openai_api_key or "").strip()
    model = (openai_model or "").strip()
    v_base = (vision_openai_api_base or "").strip()
    v_key = (vision_openai_api_key or "").strip()
    v_model = (vision_openai_model or "").strip()
    vm_raw = (vision_multimodal or "").strip().lower()
    vision_mm = vm_raw in ("1", "true", "yes", "on")
    if not api_base:
        return Response(
            content=_render_model_admin_page("保存失败：OPENAI_API_BASE 不能为空。"),
            media_type="text/html; charset=utf-8",
            status_code=400,
        )
    if not model:
        return Response(
            content=_render_model_admin_page("保存失败：OPENAI_MODEL 不能为空。"),
            media_type="text/html; charset=utf-8",
            status_code=400,
        )
    updates = {
        "OPENAI_API_BASE": api_base,
        "OPENAI_MODEL": model,
        "VISION_MULTIMODAL": vision_mm,
        "VISION_OPENAI_API_BASE": v_base,
        "VISION_OPENAI_MODEL": v_model,
    }
    if api_key:
        updates["OPENAI_API_KEY"] = api_key
    if v_key:
        updates["VISION_OPENAI_API_KEY"] = v_key
    update_runtime_config_file(updates)
    logger.info(
        "admin/model updated config.json OPENAI_API_BASE=%s OPENAI_MODEL=%s VISION_MULTIMODAL=%s VISION_OPENAI_MODEL=%s text_key=%s vision_key=%s",
        api_base,
        model,
        vision_mm,
        v_model or "(empty)",
        "yes" if api_key else "no",
        "yes" if v_key else "no",
    )
    return Response(
        content=_render_model_admin_page("保存成功：已写入 config.json 并已重新加载。"),
        media_type="text/html; charset=utf-8",
    )


@app.get("/api/funding_compare.xlsx")
def api_funding_compare_xlsx():
    """
    资金费率对比导出 Excel：拉取 Toobit / Binance 全市场对比数据，返回 .xlsx 文件供下载。
    卡片内「下载 Excel」按钮指向此接口（需配置 PUBLIC_BASE_URL 且公网可访问）。
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    from tools.funding_rate import get_funding_compare_toobit_binance

    try:
        rows = get_funding_compare_toobit_binance()
    except Exception as e:
        logger.exception("api_funding_compare_xlsx fetch error: %s", e)
        return Response(
            content=f"拉取数据失败: {e}".encode("utf-8"),
            status_code=500,
            media_type="text/plain; charset=utf-8",
        )
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return Response(content=b"internal error", status_code=500)
    ws.title = "Toobit vs 币安"
    headers = ("标的", "Toobit(%)", "币安(%)", "差值(%)")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=r.get("symbol_short", ""))
        ws.cell(row=row_idx, column=2, value=round(r.get("toobit_rate_pct", 0), 6))
        ws.cell(row=row_idx, column=3, value=round(r.get("binance_rate_pct", 0), 6))
        ws.cell(row=row_idx, column=4, value=round(r.get("diff_pct", 0), 6))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="funding_compare.xlsx"',
        },
    )


def main():
    errors = validate_webhook_config()
    if errors:
        for e in errors:
            logger.error(e)
        raise SystemExit(1)
    try:
        port = int(WEBHOOK_PORT)
    except ValueError:
        port = 9000
    logger.info("Starting Webhook server at http://%s:%s path=%s", WEBHOOK_HOST, port, WEBHOOK_PATH or "/")
    logger.info("GitLab MR Webhook URL: http://%s:%s%s", WEBHOOK_HOST, port, GITLAB_WEBHOOK_PATH)
    import uvicorn
    uvicorn.run(app, host=WEBHOOK_HOST, port=port)


if __name__ == "__main__":
    main()
